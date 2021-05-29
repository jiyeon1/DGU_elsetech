import glob
import openpyxl
from openpyxl_image_loader import SheetImageLoader
import shutil
import os

NEW_FILE_PATH = os.path.join(os.getcwd(), "get_water_info\\newData", '*.xlsx')
OLD_FILE_PATH = os.path.join(os.getcwd(), "get_water_info\\oldData")


# 엑셀 파일 읽어서 DB 저장
def read_new_data(path, new_path):
    wb = openpyxl.load_workbook(path)
    ws = wb['Sheet1']
    
    name = ws['A7'].value               # 이름
    water_origin = ws['C18'].value      # 취수원
    turbidity = ws['E18'].value         # 탁도
    fe_origin = ws['H18'].value         # 철

    cell_row = ord('D')
    cell_pos = chr(cell_row) + str(21)

    label = ['철(Fe)', '망간(Mn)', '알루미눔(Al)', '육안검사', '종합평가']
    analysis_data = []

    while ws[cell_pos].value:
        day = ws[cell_pos].value

        for i in range(5):
            # 이미지 읽기
            pos = chr(cell_row) + str(i + 22)
            if i == 3:
                image_loader = SheetImageLoader(ws)
                image = image_loader.get(pos)
                image.show()
                analysis_data.append('이미지')

                continue

            value = ws[pos].value
            analysis_data.append(value)

        print(day)

        print("이름 : " + name)

        print("- 상수도 정보- ")
        print("취수원 : " + water_origin)
        print("탁도 : " + str(turbidity))
        print("철(Fe) : " + str(fe_origin))

        print("- 수질 정보 분석 보고서 - ")
        for i in range(5):
            print(label[i] + " : " + str(analysis_data[i]))

        print("=============================")

        cell_row += 1
        cell_pos = chr(cell_row) + str(21)
        '''
        # DB에 저장
        data = AnalysisData(
            name=name,
            ...
            ...
            ...
        )
        data.save()
        '''

    # 읽은 파일 이동
    # shutil.move(path, new_path)


def read_all_data(test=False):
    global NEW_FILE_PATH, OLD_FILE_PATH

    if test:
        NEW_FILE_PATH = os.path.join(os.getcwd(), "newData", '*.xlsx')
        OLD_FILE_PATH = os.path.join(os.getcwd(), "oldData")

    path = NEW_FILE_PATH
    new_path = OLD_FILE_PATH

    file_list = glob.glob(path)
    for file in file_list:
        read_new_data(file, new_path)


# test
if __name__ == '__main__':
    read_all_data(test=True)
